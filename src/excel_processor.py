import pandas as pd
from pathlib import Path
import re

class ExcelProcessor:
    def __init__(self):
        self.input_df = None
        self.output_df = None

    def load_excel(self, file_path):
        """Load the Excel file into a pandas DataFrame."""
        try:
            print(f"\nDEBUG: Attempting to load file: {file_path}")
            print(f"DEBUG: File exists: {Path(file_path).exists()}")
            
            # List of engines to try
            engines = ['openpyxl', 'xlrd']
            
            for engine in engines:
                try:
                    print(f"DEBUG: Attempting to load with engine: {engine}")
                    # Try to read the file without header first to examine the structure
                    raw_df = pd.read_excel(file_path, engine=engine, header=None)
                    print(f"DEBUG: Successfully loaded raw Excel file with {engine}")
                    
                    # Find the actual header row by looking for key columns
                    header_row = None
                    for i in range(min(20, len(raw_df))):  # Check first 20 rows
                        row_values = raw_df.iloc[i].astype(str)
                        if any(col in row_values.values for col in ['Description', 'Message', 'Origin', 'Type']):
                            header_row = i
                            break
                    
                    # If header row found, read again with that as the header
                    if header_row is not None:
                        print(f"DEBUG: Found header at row {header_row}")
                        self.input_df = pd.read_excel(file_path, engine=engine, header=header_row)
                    else:
                        print("DEBUG: Using first row as header")
                        self.input_df = pd.read_excel(file_path, engine=engine)
                    
                    # Clean up column names
                    self.input_df.columns = [str(col).strip() for col in self.input_df.columns]
                    
                    print(f"DEBUG: DataFrame shape: {self.input_df.shape}")
                    print(f"DEBUG: Columns: {self.input_df.columns.tolist()}")
                    
                    if len(self.input_df) > 0:
                        print("\nDEBUG: File Content Preview:")
                        print(self.input_df.head())
                        return True
                    
                except Exception as e:
                    print(f"DEBUG: Error with {engine}: {str(e)}")
                    continue
            
            # Try salvaging the file when all engines fail
            print("DEBUG: All engines failed, trying direct CSV conversion...")
            
            # New fallback method: Try using a temporary conversion to CSV
            import tempfile
            import subprocess
            import os
            import csv
            
            # Create temporary CSV file
            temp_dir = tempfile.mkdtemp()
            temp_csv = os.path.join(temp_dir, "temp_excel.csv")
            
            # Use pandas direct read with errors='ignore'
            try:
                print("DEBUG: Attempting to read with pandas errors='ignore'")
                self.input_df = pd.read_excel(file_path, engine='openpyxl', header=None, errors='ignore')
                if len(self.input_df) > 0:
                    print("DEBUG: Successfully read with errors='ignore' option")
                    
                    # Find the header row by using keyword matching
                    header_row = None
                    for i in range(min(30, len(self.input_df))):  # Check more rows just in case
                        row_str = ' '.join(self.input_df.iloc[i].astype(str).values)
                        if 'Description' in row_str and ('Message' in row_str or 'Type' in row_str):
                            header_row = i
                            break
                    
                    # If header row found, use it as the header
                    if header_row is not None:
                        print(f"DEBUG: Found header at row {header_row}")
                        self.input_df.columns = self.input_df.iloc[header_row]
                        self.input_df = self.input_df.iloc[header_row + 1:]
                    
                    # Clean up column names
                    self.input_df.columns = [str(col).strip() for col in self.input_df.columns]
                    
                    print(f"DEBUG: DataFrame shape: {self.input_df.shape}")
                    print(f"DEBUG: Columns: {self.input_df.columns.tolist()}")
                    
                    return True
            except Exception as e:
                print(f"DEBUG: Error with pandas errors='ignore': {str(e)}")
            
            # Try a manual parsing approach
            print("DEBUG: Trying manual parsing approach...")
            try:
                # Create a manually parsed dataframe
                import openpyxl
                from openpyxl.utils.exceptions import InvalidFileException
                
                try:
                    # Try a more lenient approach with openpyxl
                    wb = openpyxl.load_workbook(file_path, data_only=True, keep_links=False, read_only=True)
                    print(f"DEBUG: Available worksheets: {wb.sheetnames}")
                    
                    if wb.sheetnames:
                        ws = wb[wb.sheetnames[0]]
                        
                        # Extract data from worksheet
                        data = []
                        for row in ws.rows:
                            row_data = [cell.value for cell in row]
                            data.append(row_data)
                        
                        if data:
                            # Find the header row
                            header_row = None
                            for i, row in enumerate(data[:30]):  # Check first 30 rows
                                row_str = ' '.join([str(cell) for cell in row if cell])
                                if row_str and ('Description' in row_str) and ('Message' in row_str or 'Type' in row_str):
                                    header_row = i
                                    break
                            
                            # Create the DataFrame
                            if header_row is not None:
                                self.input_df = pd.DataFrame(data[header_row+1:], columns=data[header_row])
                            else:
                                self.input_df = pd.DataFrame(data[1:], columns=data[0])
                            
                            # Clean up column names
                            self.input_df.columns = [str(col).strip() if col else f"Column_{i}" for i, col in enumerate(self.input_df.columns)]
                            
                            # Drop empty columns
                            self.input_df = self.input_df.loc[:, ~self.input_df.columns.str.contains('^Column_')]
                            
                            print(f"DEBUG: Successfully created DataFrame with shape: {self.input_df.shape}")
                            print(f"DEBUG: Columns: {self.input_df.columns.tolist()}")
                            
                            if len(self.input_df) > 0:
                                print("\nDEBUG: File Content Preview:")
                                print(self.input_df.head())
                                return True
                except InvalidFileException:
                    print("DEBUG: InvalidFileException with openpyxl")
                except Exception as e:
                    print(f"DEBUG: Error with openpyxl manual parsing: {str(e)}")
            
            except Exception as e:
                print(f"DEBUG: Error with manual parsing: {str(e)}")
            
            # If all else fails
            print("DEBUG: All methods failed. Recommending to repair the file.")
            print("DEBUG: Try opening and saving the file with Microsoft Excel, Google Sheets, or LibreOffice Calc.")
            return False
            
        except Exception as e:
            print(f"\nDEBUG: Error details:")
            print(f"- Error type: {type(e).__name__}")
            print(f"- Error message: {str(e)}")
            print(f"- File path: {file_path}")
            return False

    def process_file(self, columns_to_translate=None):
        """Process the loaded Excel file."""
        if self.input_df is None:
            print("DEBUG: input_df is None")
            return False
        
        if columns_to_translate is None:
            columns_to_translate = ["Description"]
            
        print(f"\nDEBUG: Starting file processing with columns to translate: {columns_to_translate}")
        
        # Create a copy of the input DataFrame
        self.output_df = self.input_df.copy()
        print(f"DEBUG: Created output DataFrame with {len(self.output_df)} rows")
        
        # Check if the specified columns exist in the DataFrame
        # First get a normalized list of available columns (removing Unnamed ones)
        available_columns = [col for col in self.output_df.columns if 'Unnamed' not in str(col)]
        print(f"DEBUG: Available columns for translation: {available_columns}")
        
        # Check for exact matches first
        missing_columns = [col for col in columns_to_translate if col not in self.output_df.columns]
        
        # If there are missing columns, try case-insensitive matching
        if missing_columns:
            column_mapping = {}
            lower_columns = {str(col).lower(): col for col in self.output_df.columns}
            
            for col in missing_columns[:]:  # Use a copy since we'll modify the list
                if col.lower() in lower_columns:
                    # Found a case-insensitive match
                    actual_col = lower_columns[col.lower()]
                    column_mapping[col] = actual_col
                    missing_columns.remove(col)
                    print(f"DEBUG: Found case-insensitive match for '{col}': '{actual_col}'")
            
            # Update columns_to_translate with the actual column names
            columns_to_translate = [column_mapping.get(col, col) for col in columns_to_translate]
        
        # Check if any columns are still missing
        missing_columns = [col for col in columns_to_translate if col not in self.output_df.columns]
        if missing_columns:
            print(f"DEBUG: Columns not found: {missing_columns}")
            print(f"DEBUG: Available columns: {self.output_df.columns.tolist()}")
            return False
            
        try:
            # Technical terms translation dictionary (English to French)
            translations = {
                # Existing translations
                'ABSENCE OF REFERENCE VOLTAGE': 'ABSENCE DE TENSION DE REFERENCE',
                'ABSENCE OF VOLTAGE': 'ABSENCE DE TENSION',
                'DEAD INCOMING DEAD RUNNING': 'ENTRÉE HORS TENSION, FONCTIONNEMENT HORS TENSION',  # Fixed spacing
                'DEAD INCOMING  DEAD RUNNING': 'ENTRÉE HORS TENSION, FONCTIONNEMENT HORS TENSION',  # Extra space variant
                'LIVE INCOMING LIVE RUNNING': 'ENTRÉE SOUS TENSION, FONCTIONNEMENT SOUS TENSION',   # Fixed spacing
                'LIVE INCOMING  LIVE RUNNING': 'ENTRÉE SOUS TENSION, FONCTIONNEMENT SOUS TENSION',  # Extra space variant
                'CLOSE PERMISSIVE': 'AUTORISATION DE FERMETURE',
                
                # Adding basic translations for Message column values
                'Set': 'Réglé',
                'Set - App Ack': 'Réglé - App Ack',
                'SET': 'RÉGLÉ',
                'SET - APP ACK': 'RÉGLÉ - APP ACK',
                'Reset': 'Réinitialiser',
                'Reset - App Ack': 'Réinitialiser - App Ack',
                'RESET': 'RÉINITIALISER',
                'RESET - APP ACK': 'RÉINITIALISER - APP ACK',
                'Operational': 'Opérationnel',
                'OPERATIONAL': 'OPÉRATIONNEL',
                'Alarm': 'Alarme',
                'ALARM': 'ALARME',
                'Normal': 'Normal',
                'NORMAL': 'NORMAL',
                'Operated': 'Opéré',
                'OPERATED': 'OPÉRÉ',
                
                # New translations
                'PRESENE OF REFERENCE VOLTAGE': 'PRÉSENCE DE TENSION DE RÉFÉRENCE',
                'PRASENCE OF VOLTAGE': 'PRÉSENCE DE TENSION',
                'LIVE INCOMING  DEAD RUNNING': 'ENTRÉE SOUS TENSION, FONCTIONNEMENT HORS TENSION',
                'POSSIBLE CLOSING': 'FERMETURE AUTORISEE',
                'BUS-1 SELECT': 'SÉLECTION DU BUS-1',
                'BUS-2 DESELECT': 'DÉSÉLECTION DU BUS-2',  # New translation
                'BAY L/R MODE': 'MODE LOCAL/DISTANT DE LA TRAVÉE',
                'BAY L/R  MODE': 'MODE L/R TRAVEE',  # Extra space variant
                'IINTERLOCK PERMISSIVE': 'VERROUILLAGE AUTORISÉ',
                'CARRIER IN': 'PORTEUSE ENTRANTE',
                'CARRIER OUT': 'PORTEUSE SORTANTE',
                'EQUIPMENT BCU': 'EQUIPEMENT BCU',  # Bloc de Contrôle Unité
                'COMMUNICATION': 'COMMUNICATION',
                'UNLOCKING RELAY ACTIVATED': 'RELAIS DEVERROUILLAGE ACTIVE',
                'PROTECTION': 'PROTECTION',
                'STAGE START': 'DEMARRAGE ETAPE',
                'STAGE': 'ETAPE',
                'SEND CH-1': 'ENVOI CANAL 1',
                'SEND CH-2': 'ENVOI CANAL 2',
                'ZONE PROTECTION': 'PROTECTION DE ZONE',
                'BPH PROTECTION': 'PROTECTION BPH',  # Bus Phase Protection Haute
                'RPH PROTECTION': 'PROTECTION RPH',  # Remote Phase Protection Haute
                'YPH PROTECTION': 'PROTECTION YPH',  # Yard Phase Protection Haute
                
                # Overcurrent protection stages
                '50/51 STAGE-1': 'ETAPE-1 50/51',  # Overcurrent protection stage 1
                '50/51 STAGE-1 START': 'DEMARRAGE ETAPE-1 50/51',
                '50/51 STAGE-2': 'ETAPE-2 50/51',
                '50/51 STAGE-2 START': 'DEMARRAGE ETAPE-2 50/51',
                
                # Communication
                'DT SEND CH-1': 'ENVOI CANAL-1 DT',
                'DT SEND CH-2': 'ENVOI CANAL-2 DT',
                
                # Directional Protection
                '67N STAGE-1': 'ETAPE-1 67N',
                '67N STAGE-1 START': 'DEMARRAGE ETAPE-1 67N',
                
                # Zone Protection
                '21 ZONE-1 PROTECTION': 'PROTECTION ZONE-1 21',
                '21 ZONE-2 PROTECTION': 'PROTECTION ZONE-2 21',
                '21 ZONE-3 PROTECTION': 'PROTECTION ZONE-3 21',
                '21 ZONE-4 PROTECTION': 'PROTECTION ZONE-4 21',
                '21 ZONE-1 PROTECTION START': 'DEMARRAGE PROTECTION ZONE-1 21',
                '21 ZONE-2 PROTECTION START': 'DEMARRAGE PROTECTION ZONE-2 21',
                '21 ZONE-3 PROTECTION START': 'DEMARRAGE PROTECTION ZONE-3 21',
                '21 ZONE-4 PROTECTION START': 'DEMARRAGE PROTECTION ZONE-4 21',
                '21 ZONE-1 YPH PROTECTION': 'PROTECTION YPH ZONE-1 21',
                '21 ZONE-1 BPH PROTECTION': 'PROTECTION BPH ZONE-1 21',
                
                # Differential Protection
                '87L PROTECTION': 'PROTECTION 87L',
                '87L PROTECTION START': 'DEMARRAGE PROTECTION 87L',
                '87L PROTECTION A-PH': 'PROTECTION 87L PHASE-A',
                
                # Rest of existing translations
                'ON/OFF SECONDARY SPS': 'MARCHE/ARRET SPS SECONDAIRE',
                'ON/OFF MAIN FOR CB1': 'MARCHE/ARRET PRINCIPAL POUR CB1',
                'DUMMY': 'FACTICE/RESERVE',
                'COMP. POSITION': 'POSITION COMP.',
                'COMP_POSITION': 'POSITION_COMP',
                'DISCONNECTOR G1 POSITION': 'POSITION SECTIONNEUR G1',
                'DISCONNECTOR G2 POSITION': 'POSITION SECTIONNEUR G2',
                'DISCONNECTOR G3 POSITION': 'POSITION SECTIONNEUR G3',
                'TRIP CIRCUIT FAULT': 'DEFAUT CIRCUIT DE DECLENCHEMENT',
                'I/L PERMISSIVE': 'PERMISSIF V/F',
                'CLOSE I/L PERMISSIVE': 'PERMISSIF V/F FERMETURE',
                'OPEN I/L PERMISSIVE': 'PERMISSIF V/F OUVERTURE',
                'CIRCUIT BREAKER GCB1 POSITION': 'DISJONCTEUR GCB1 POSITION',
                'CIRCUIT BREAKER-GCB1 POS': 'DISJONCTEUR-GCB1 POS',
                'BUS-1 DESELECT': 'DÉSÉLECTION JEU DE BARRES-1',
                'CB CLOSE ORDER': 'ORDRE DE FERMETURE DISJONCTEUR',
                'ORDER RUNNING': 'ORDRE EN COURS',
                'INTERLOCK PERMISSIVE': 'VERROUILLAGE AUTORISÉ',
                'OPERATE': 'OPÉRER',
                'SELECT': 'SÉLECTIONNER',
                'SYNCHROCHECK IN PROGRESS': 'VÉRIFICATION SYNCHRO EN COURS',
                'GENERAL TRIP': 'DÉCLENCHEMENT GÉNÉRAL',
                '27 STAGE-1 START': 'DÉMARRAGE ÉTAPE-1 27',  # Protection minimum de tension
                '27 STAGE-2': 'ÉTAPE-2 27',
                '50N/51N OPTD': '50N/51N OPÉRÉ',  # Protection à maximum de courant terre
                'OPERATING MODE': 'MODE DE FONCTIONNEMENT',
                '21 ZONE-1 C-PH OPTD': '21 ZONE-1 PHASE-C OPÉRÉE',  # Protection de distance
                'CARRIER SEND CHANNEL-1': 'ENVOI PORTEUSE CANAL-1',
                '24 ALARM': 'ALARME 24',  # Protection de surexcitation V/Hz
                'HV 64REF': 'PROTECTION TERRE RESTREINTE 64 HT',
                '2ND HARMONIC DETECTED': '2ÈME HARMONIQUE DÉTECTÉ',
                '87T C-PH OPTD': '87T PHASE-C OPÉRÉE',  # Protection différentielle transformateur
                'TIME SYNCHRONISATION': 'SYNCHRONISATION TEMPORELLE',
                '24 STAGE-1 START': 'DÉMARRAGE ÉTAPE-1 24',

                # New transformer differential protection translations
                '87T A-PH OPTD': '87T PHASE-A OPÉRÉE',
                '87T B-PH OPTD': '87T PHASE-B OPÉRÉE',
                '87T OPTD': '87T OPÉRÉE',
                
                # High voltage earth fault protection translations
                'HV 50N/51N STAGE-1 START': 'DÉMARRAGE ÉTAPE-1 50N/51N HT',
                'HV 50N/51N STAGE-1': 'ÉTAPE-1 50N/51N HT',
                'HV 50N/51N STAGE-2 START': 'DÉMARRAGE ÉTAPE-2 50N/51N HT',  # Protection à maximum de courant terre haute tension
                'HV 50N/51N STAGE-2': 'ÉTAPE-2 50N/51N HT',
                
                # Overcurrent protection translations
                '50/51 STAGE-1 A-PH': 'ÉTAPE-1 50/51 PHASE-A',
                '50/51 STAGE-1 B-PH': 'ÉTAPE-1 50/51 PHASE-B',  # Protection à maximum de courant phase B
                '50/51 STAGE-1 C-PH': 'ÉTAPE-1 50/51 PHASE-C',  # Protection à maximum de courant phase C
                '24 STAGE-1': 'ÉTAPE-1 24',  # Protection de surexcitation V/Hz

                # Switchgear and operational status translations
                '+SWG EFS B8 OPERATIONAL': '+TBT EFS B8 OPÉRATIONNEL',
                '+6R3 EFS B2 OPERATIONAL': '+6R3 EFS B2 OPÉRATIONNEL',
                '+6R3 EFS B5 OPERATIONAL': '+6R3 EFS B5 OPÉRATIONNEL',
                '+SWG EFS B7 OPERATIONAL': '+TBT EFS B7 OPÉRATIONNEL',
                
                # DC circuit breaker translations
                'DC MCB TRIP': 'DÉCLENCHEMENT DISJONCTEUR CC',
                '6MET-DC MCB TRIP': 'DÉCLENCHEMENT DISJONCTEUR CC 6MET',

                # New translations
                'BAY MODE': 'MODE TRAVÉE',
                'MODE TRAVEL': 'MODE TRAVÉE',
                '+6R3 EFS B3 OPERATIONAL': '+6R3 EFS B3 EN SERVICE',
                '+6R1 EFS B1 OPERATIONAL': '+6R1 EFS B1 EN SERVICE',
                '+6R3 EFS B4 OPERATIONAL': '+6R3 EFS B4 EN SERVICE',
                'REGULATOR R/L': 'RÉGULATEUR D/G',
                'MOTOR MCB FAIL': 'DÉFAUT DISJONCTEUR MOTEUR',
                'TAP CHANGER IN SERVICE': 'CHANGEUR DE PRISES EN SERVICE',
                '21 OPTD': '21 DÉCLENCHÉE',
                '67N OPTD': '67N DÉCLENCHÉE',
                '50/51 OPTD': '50/51 DÉCLENCHÉE',
                '81 OF STAGE-1': '81 OF SEUIL-1',
                '21 ZONE-1 B-PH OPTD': '21 ZONE-1 PHASE-B DÉCLENCHÉE',
                '21 ZONE-1 START': '21 ZONE-1 DÉMARRAGE',
                '21 ZONE-4 START': '21 ZONE-4 DÉMARRAGE',
                '81UF STAGE-1 START': '81UF SEUIL-1 DÉMARRAGE',
                '81 UF STAGE-1': '81 UF SEUIL-1',
                '81OF STAGE-1 START': '81OF SEUIL-1 DÉMARRAGE',
                '27 STAGE-1': '27 SEUIL-1',
                '59 STAGE-1 START': '59 SEUIL-1 DÉMARRAGE',
                '59 STAGE-2': '59 SEUIL-2',
                '59 STAGE-1': '59 SEUIL-1',
                '67 OPTD': '67 DÉCLENCHÉE',
                '21 ZONE-1 PROTECTION OPTD': '21 ZONE-1 PROTECTION DÉCLENCHÉE',
                '21 ZONE-1 A-PH OPTD': '21 ZONE-1 PHASE-A DÉCLENCHÉE',
                '21 ZONE-4 PROTECTION': '21 ZONE-4 PROTECTION',
                '21 ZONE-3 START': '21 ZONE-3 DÉMARRAGE',
                '21 ZONE-2 START': '21 ZONE-2 DÉMARRAGE',
                '21 ZONE-2 PROTECTION OPTD': '21 ZONE-2 PROTECTION DÉCLENCHÉE',
                
                # Message column specific translations
                'REMOTE': 'DISTANT',
                'BAD STATE': 'MAUVAIS ÉTAT',
                'OPEN': 'OUVERT',
                'ON': 'ACTIVÉ',
                'Set': 'Réglé',
                'Set   ': 'Réglé',
                'SET': 'RÉGLÉ',
                'Reset': 'Réinitialisé',
                'RESET': 'RÉINITIALISÉ',
                'Reset - App Ack': 'Réinitialisé - App Ack',
                'Reset - App Ack   ': 'Réinitialisé - App Ack',
                'RESET - APP ACK': 'RÉINITIALISÉ - APP ACK',
                'OPEN - App Ack': 'OUVERT - App Ack',
                'OPEN - App Ack   ': 'OUVERT - App Ack',
                'OPEN - Clearing': 'OUVERT - Effacement',
                'OPEN - CLEARING': 'OUVERT - EFFACEMENT',
                'Set - App Ack': 'Réglé - App Ack',
                'Set - App Ack   ': 'Réglé - App Ack',
                'SET - APP ACK': 'RÉGLÉ - APP ACK',
                'Set - Clearing': 'Réglé - Effacement',
                'Set - Clearing   ': 'Réglé - Effacement',
                'SET - CLEARING': 'RÉGLÉ - EFFACEMENT',
                'On Sync': 'En Synchronisation',
                'ON SYNC': 'EN SYNCHRONISATION',
                'TRIP - App Ack': 'DÉCLENCHEMENT - App Ack',
                'TRIP - APP ACK': 'DÉCLENCHEMENT - APP ACK',
                'TRIP': 'DÉCLENCHEMENT',
                'Trip': 'Déclenchement',
                'Closed': 'Fermé',
                'CLOSED': 'FERMÉ',
                'Operated': 'Opéré',
                'OPERATED': 'OPÉRÉ',
                'Operated - Clearing': 'Opéré - Effacement',
                'Operated - Clearing   ': 'Opéré - Effacement',
                'OPERATED - CLEARING': 'OPÉRÉ - EFFACEMENT',
                'Operated - App Ack': 'Opéré - App Ack',
                'Operated - App Ack   ': 'Opéré - App Ack',
                'OPERATED - APP ACK': 'OPÉRÉ - APP ACK',
                'Healthy': 'En Bon État',
                'HEALTHY': 'EN BON ÉTAT',
                'Fail': 'Défaillance',
                'FAIL': 'DÉFAILLANCE',
                'Faulty': 'Défectueux', 
                'FAULTY': 'DÉFECTUEUX',
                'Faulty - Clearing': 'Défectueux - Effacement',
                'Faulty - Clearing   ': 'Défectueux - Effacement',
                'FAULTY - CLEARING': 'DÉFECTUEUX - EFFACEMENT',
                'Fail - Clearing': 'Défaillance - Effacement',
                'Fail - Clearing   ': 'Défaillance - Effacement',
                'FAIL - CLEARING': 'DÉFAILLANCE - EFFACEMENT',
                'Fail - App Ack': 'Défaillance - App Ack',
                'Fail - App Ack   ': 'Défaillance - App Ack',
                'FAIL - APP ACK': 'DÉFAILLANCE - APP ACK',
                'ABSENCE TENSION': 'ABSENCE DE TENSION',
                'DEFAULT ALIM CG MCB1/MCB2 DECLENCHEE': 'DÉFAUT ALIM CG MCB1/MCB2 DÉCLENCHÉE',
                'EFS-52 OPERATIONAL': 'EFS-52 OPÉRATIONNEL',
                'EFS-SB2 OPERATIONAL': 'EFS-SB2 OPÉRATIONNEL',
                'ABSENCE TENSION 125V CG2': 'ABSENCE DE TENSION 125V CG2',
                'DISJONCTEUR QE1': 'DISJONCTEUR QE1',
                'DISJONCTEUR QE2': 'DISJONCTEUR QE2',
                'DISJONCTEUR QE3': 'DISJONCTEUR QE3',
                
                # From image
                'DISJONCTEUR QR3': 'DISJONCTEUR QR3',
                'DISJONCTEUR QR4': 'DISJONCTEUR QR4', 
                'DISJONCTEUR QB1': 'DISJONCTEUR QB1',
                'DISJONCTEUR QS1': 'DISJONCTEUR QS1',
                'DISJONCTEUR QS2': 'DISJONCTEUR QS2',
                'DISJONCTEUR QQ2': 'DISJONCTEUR QQ2',
                'DISJONCTEUR QG1': 'DISJONCTEUR QG1',
                'DISJONCTEUR QD1': 'DISJONCTEUR QD1',
                'BAD STATE': 'MAUVAIS ÉTAT',
                'EN SERVICE': 'EN SERVICE',
                'OPÉRATIONNEL': 'OPÉRATIONNEL',
                'Operational Mode': 'Mode Opérationnel',
                'OPERATIONAL MODE': 'MODE OPÉRATIONNEL',
                'OFF': 'DÉSACTIVÉ',
                'Off': 'Désactivé',
                'OFF - App Ack': 'DÉSACTIVÉ - App Ack',
                'OFF - App Ack   ': 'DÉSACTIVÉ - App Ack',
                'Off - App Ack': 'Désactivé - App Ack',
                'Off - App Ack   ': 'Désactivé - App Ack',
                'Alarm': 'Alarme',
                'ALARM': 'ALARME'
            }
            
            def is_french(text):
                """Check if text contains French-specific words/patterns"""
                french_indicators = [
                    'DE', 'DES', 'PERMISSIF', 'SECTIONNEUR', 'TERRE', 'DISJONCTEUR', 
                    'MARCHE', 'ARRET', 'ENTREE', 'INACTIVE', 'EXECUTION', 'COMMANDE',
                    'MANUELLE', 'PORTEUSE', 'ENTRANTE', 'SORTANTE', 'EQUIPEMENT',
                    'RELAIS', 'DEVERROUILLAGE', 'DEMARRAGE', 'ETAPE', 'ENVOI', 'CANAL',
                    'PROTECTION', 'PHASE', 'ZONE', 'MAUVAIS', 'ÉTAT', 'OUVERT', 'ACTIVÉ'
                ]
                
                english_indicators = [
                    'OPERATING', 'MODE', 'HARMONIC', 'DETECTED', '2ND', 'BAY', 'REMOTE', 'OPERATIONAL'
                ]
                
                words = text.upper().split()
                
                # If any word is explicitly English, return False
                if any(word in english_indicators for word in words):
                    return False
                
                french_word_count = sum(1 for word in words if any(indicator == word for indicator in french_indicators))
                total_words = len(words)
                
                return (french_word_count / total_words) > 0.3 if total_words > 0 else False
            
            def translate_text(text):
                if pd.isna(text) or text is None or str(text).strip() == '':
                    return text
                    
                # Handle specific cases with trailing spaces
                original_text = str(text)
                stripped_text = original_text.strip()
                
                special_cases = {
                    "Set - App Ack": "Réglé - App Ack",
                    "Reset - App Ack": "Réinitialisé - App Ack",
                    "TRIP - App Ack": "DÉCLENCHEMENT - App Ack",
                    "OPEN - App Ack": "OUVERT - App Ack",
                    "OPEN - Clearing": "OUVERT - Effacement",
                    "Set - Clearing": "Réglé - Effacement",
                    "Fail - App Ack": "Défaillance - App Ack",
                    "Fail - Clearing": "Défaillance - Effacement",
                    "Faulty - Clearing": "Défectueux - Effacement",
                    "Operated - Clearing": "Opéré - Effacement",
                    "Operated - App Ack": "Opéré - App Ack",
                    "OFF - App Ack": "DÉSACTIVÉ - App Ack",
                    "Off - App Ack": "Désactivé - App Ack",
                    "Operated": "Opéré",
                    "Trip": "Déclenchement",
                    "Closed": "Fermé",
                    "On Sync": "En Synchronisation",
                    "Healthy": "En Bon État",
                    "Operational Mode": "Mode Opérationnel",
                    "Alarm": "Alarme",
                    "Fail": "Défaillance",
                    "Faulty": "Défectueux",
                    "Off": "Désactivé"
                }
                
                # Check common patterns with trailing spaces that cause issues
                known_patterns = {
                    "Set - App Ack": "Réglé - App Ack",
                    "Reset - App Ack": "Réinitialisé - App Ack",
                    "OPEN - App Ack": "OUVERT - App Ack",
                    "OPEN - Clearing": "OUVERT - Effacement",
                    "Set - Clearing": "Réglé - Effacement",
                    "Fail - App Ack": "Défaillance - App Ack",
                    "Fail - Clearing": "Défaillance - Effacement",
                    "Faulty - Clearing": "Défectueux - Effacement",
                    "Operated - Clearing": "Opéré - Effacement",
                    "Operated - App Ack": "Opéré - App Ack",
                    "OFF - App Ack": "DÉSACTIVÉ - App Ack",
                    "Off - App Ack": "Désactivé - App Ack",
                    "Operational Mode": "Mode Opérationnel",
                    "Set": "Réglé",
                    "Reset": "Réinitialisé",
                    "Operated": "Opéré",
                    "Off": "Désactivé",
                    "Alarm": "Alarme"
                }
                
                # Check for exact matches including spaces first (more specific)
                for pattern, translation in known_patterns.items():
                    # If the text starts with the pattern plus optional spaces, translate it
                    if re.match(f"^{re.escape(pattern)}\\s*$", original_text, re.IGNORECASE):
                        print(f"DEBUG: Translated '{text}' → '{translation}' (pattern with spaces)")
                        return translation
                
                # Check if the stripped text matches any of our special cases (more general)
                for case, translation in special_cases.items():
                    if stripped_text.upper() == case.upper() or stripped_text.upper().startswith(case.upper() + " "):
                        print(f"DEBUG: Translated '{text}' → '{translation}' (special case)")
                        return translation
                    
                # Normalize the text by removing extra spaces (both within and at the end)
                text_str = ' '.join(str(text).strip().upper().split())
                
                # First check if the text is French
                if is_french(text_str):
                    print(f"DEBUG: Keeping French text: '{text}'")
                    return text
                
                # If not French, try to translate from English to French
                if text_str in translations:
                    translated = translations[text_str]
                    print(f"DEBUG: Translated '{text}' → '{translated}'")
                    return translated
                
                # Try with original spacing if normalized version not found
                original_upper = str(text).strip().upper()
                if original_upper in translations:
                    translated = translations[original_upper]
                    print(f"DEBUG: Translated '{text}' → '{translated}'")
                    return translated
                    
                # Try with additional variations for handling whitespace
                for key in translations:
                    # Try matching ignoring extra spaces
                    if ' '.join(text_str.split()) == ' '.join(key.split()):
                        translated = translations[key]
                        print(f"DEBUG: Translated '{text}' → '{translated}' (space-normalized)")
                        return translated
                
                # If no translation found, return original
                print(f"DEBUG: No translation found for '{text}', keeping original")
                return text
            
            # Apply translation to each selected column
            for column in columns_to_translate:
                print(f"\nDEBUG: Starting translation of '{column}' column")
                
                # Create a new column for the translation
                new_column_name = f"{column} Français"
                
                # Translate the column and add it next to the original column
                translated_values = self.output_df[column].apply(translate_text)
                
                # Get the position of the current column
                column_position = self.output_df.columns.get_loc(column)
                
                # Create a new DataFrame with all columns before the current one
                columns_before = list(self.output_df.columns[:column_position + 1])
                
                # Create a list of all columns after the current one
                columns_after = list(self.output_df.columns[column_position + 1:])
                
                # Reorganize the DataFrame
                self.output_df = pd.concat([
                    self.output_df[columns_before], 
                    translated_values.rename(new_column_name),
                    self.output_df[columns_after]
                ], axis=1)
                
                print(f"DEBUG: Added new column '{new_column_name}'")
            
            print("DEBUG: Translation completed")
            
            # Limit to 1050 rows if necessary
            if len(self.output_df) > 1050:
                self.output_df = self.output_df.head(1050)
                print("DEBUG: Truncated to 1050 rows")
            
            return True
            
        except Exception as e:
            print(f"\nDEBUG: Error during translation:")
            print(f"- Error type: {type(e).__name__}")
            print(f"- Error message: {str(e)}")
            return False

    def save_excel(self, output_path):
        """Save the processed DataFrame to a new Excel file."""
        if self.output_df is None:
            print("DEBUG: output_df is None")
            return False
            
        try:
            print(f"\nDEBUG: Attempting to save file to: {output_path}")
            print(f"DEBUG: DataFrame shape: {self.output_df.shape}")
            self.output_df.to_excel(output_path, index=False, engine='openpyxl')
            print("DEBUG: File saved successfully")
            return True
        except Exception as e:
            print(f"\nDEBUG: Error while saving:")
            print(f"- Error type: {type(e).__name__}")
            print(f"- Error message: {str(e)}")
            return False 